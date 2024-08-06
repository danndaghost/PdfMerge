import os
import time
import yaml
from openpyxl import load_workbook
from fnmatch import fnmatch
from PyPDF2 import PdfReader, PdfWriter

#variables globales
rootpath = ''
writepath = ''
readfile = ''
readpath = []
excludepath = []
allfiles = []
pdffiles = {}
config = {}
dotacion = []
columnignore = 0
columnrut = 0

def merge_pdfs(paths, output):
    pdf_writer = PdfWriter()
    for path in paths:
        pdf_reader = PdfReader(path)
        for page in range(len(pdf_reader.pages)):
            # Add each page to the writer object
            pdf_writer.add_page(pdf_reader.pages[page])
    # Write out the merged PDF
    with open(output, 'wb') as out:
        pdf_writer.write(out)

def leerExcel():
    filepath        = rootpath
    archivo         = filepath + '\\' + writepath + '\\'   
    archivo_excel   = archivo + readfile #la extension xlsx esta en la configuracion
    try:
        workbook = load_workbook(archivo_excel)
        hoja = workbook.worksheets[0]
        for row in range(2, hoja.max_row+1): #start on two to ignore the title of column
            if(hoja.cell(row=row, column=columnignore).value != 'SI'):
                dotacion.append(hoja.cell(row=row, column=columnrut).value)
        return dotacion
    except FileNotFoundError as exc:
        print("Archivo no encontrado, verifique la ruta y si está conectado a la VPN")
        exit()

def leerArchivos():
    for rootfolder in readpath:
        for path, dirs, files in os.walk(rootpath + '\\' + rootfolder):
            dirs[:] = [d for d in dirs if d not in excludepath]
            for name in files:
                if fnmatch(name, "*.pdf"): #considerar solo archivos PDF
                    allfiles.append(os.path.join(path, name))
    return allfiles

def leerRutArchivo(allfiles, dotacion):
    for rut in dotacion:
        pdffiles[rut] = []
        for file in allfiles:
            if rut in file:
                pdffiles[rut].append(file)
    return pdffiles

def crearPdf(pdffiles):
    contador = 0
    #obtener la cantidad de keys    
    rutArchivo = pdffiles.keys()
    cantidadtotal = len(rutArchivo)
    for rut in rutArchivo:
        contador = contador+1
        #print(":::: " + str(contador) + " / " + str(cantidadtotal) +  "   ::::   RUT  " + rut + " ::::")
        print(":::: EN PROCESO {:2.1%}".format(contador / cantidadtotal), end="\r")
        merge_pdfs(pdffiles[rut], output=rootpath + '\\' + writepath + '\\' + rut + '.penalolen.pdf')    

def config():
    with open("config.yml") as stream:
        try:
            config = yaml.safe_load(stream)

            global rootpath
            rootpath = config['ruta_lectura']

            global readpath
            readpath = config['carpeta_lectura']

            global excludepath
            excludepath = config['carpeta_no_lectura']

            global writepath
            writepath = config['carpeta_escritura']

            global readfile
            readfile = config['archivo_dotacion']

            global columnrut
            columnrut = config['columna_rut']

            global columnignore
            columnignore = config['columna_ignorar']

        except yaml.YAMLError as exc:
            print(exc)
            exit()        

if __name__ == "__main__":
    config()

    print(":::: PROCESO INICIADO A LAS " + time.strftime("%H:%M") + " ::::")
    #Obtener los rut desde archivo excel
    print(":::: OBTENIENDO DOTACION ACTUALIZADA ::::")
    dotacion = leerExcel() 
    print(":::: SE ENCONTRARON " + str(len(dotacion)) + " REGISTROS ::::")

    #Obtener el listado de archivos del servidor
    print(":::: LEYENDO CARPETAS COMPARTIDAS ::::")
    allfiles = leerArchivos()
    print(":::: SE ENCONTRARON " + str(len(allfiles)) + " ARCHIVOS ::::")

    #Obtener el emparejamiento archivo / rut
    print(":::: BUSCANDO RUT EN CARPETAS COMPARTIDAS ::::")
    pdffiles = leerRutArchivo(allfiles, dotacion)
    
    #Crear PDF unificado de cada RUT
    print(":::: GENERANDO PDF UNIFICADOS ::::")
    crearPdf(pdffiles)

    print(":::: PROCESO FINALIZADO A LAS " + time.strftime("%H:%M") + " ::::")
   